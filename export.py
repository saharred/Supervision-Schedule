"""
Export functionality for supervision schedule
Supports Excel and PDF export
"""

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from datetime import datetime
import io


def export_to_excel(df, filename="supervision_schedule.xlsx"):
    """Export DataFrame to Excel with formatting"""
    if df.empty:
        return None
    
    # Create Excel writer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='جدول المراقبة', index=False, encoding='utf-8-sig')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['جدول المراقبة']
        
        # Set RTL
        worksheet.sheet_view.rightToLeft = True
        
        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 5, 50)
        
        # Format header
        from openpyxl.styles import Font, Alignment, PatternFill
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    output.seek(0)
    return output


def export_to_pdf(df, school_name="مدرسة عثمان بن عفان", filename="supervision_schedule.pdf"):
    """Export DataFrame to PDF with Arabic support"""
    if df.empty:
        return None
    
    output = io.BytesIO()
    
    # Create PDF
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=50,
        bottomMargin=30
    )
    
    # Register Arabic font (using default for now, can be enhanced)
    try:
        # Try to use system Arabic font
        pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        arabic_font = 'Arabic'
    except:
        arabic_font = 'Helvetica'
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ArabicTitle',
        parent=styles['Heading1'],
        fontName=arabic_font,
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    subtitle_style = ParagraphStyle(
        'ArabicSubtitle',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    # Build content
    content = []
    
    # Title
    title = Paragraph(f"<b>{school_name}</b>", title_style)
    content.append(title)
    
    subtitle = Paragraph("جدول المراقبة للاختبارات", subtitle_style)
    content.append(subtitle)
    
    date_text = Paragraph(f"تاريخ الإصدار: {datetime.now().strftime('%Y-%m-%d')}", subtitle_style)
    content.append(date_text)
    
    content.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    table_data = [df.columns.tolist()] + df.values.tolist()
    
    # Create table
    table = Table(table_data, repeatRows=1)
    
    # Table style
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTNAME', (0, 1), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Data rows
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    content.append(table)
    
    # Build PDF
    doc.build(content)
    
    output.seek(0)
    return output

